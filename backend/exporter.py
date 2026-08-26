import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Import the calculator logic so the loop can use it
from backend.calculator import depreciation_for_row

def make_excel(df, header, rc_data, metal_rate, doc_type, age_text):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assessment"
    ws["A1"] = "MOTOR CLAIM AI REPORT SYSTEM - INVOICE ASSESSMENT"
    ws["A1"].font = Font(bold=True, size=14)
    
    # Fixed merge area to cover all 15 columns (A through O)
    ws.merge_cells("A1:O1") 
    
    fields = [
        ("Invoice No.", header.get("Invoice No.","")),
        ("Invoice Date", header.get("Invoice Date","")),
        ("Document Type", doc_type),  # Passed as argument now
        ("Workshop", header.get("Workshop / Supplier","")),
        ("Registration No.", rc_data.get("Registration No.") or header.get("Registration No.","")),
        ("Owner", rc_data.get("Owner / Customer") or header.get("Owner / Customer","")),
        ("Registration Date", rc_data.get("Registration Date","")),
        ("Vehicle Age", age_text),    # Passed as argument now
        ("Metal Depreciation", f"{metal_rate}%"),
    ]
    
    r = 3
    for k, v in fields:
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
        r += 1
        
    start = r + 1
    cols = ["Select", "Item Type", "Description", "Part No. / Labour Code", "HSN / SAC", "PMG", "Qty", "Unit", "Rate (₹)", "Taxable Amount (₹)", "GST %", "GST Amount (₹)", "Dep. %", "Dep. Amount (₹)", "Net After Dep. (₹)"]
    
    for c, h in enumerate(cols, 1):
        cell = ws.cell(start, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        
    for i, (_, row) in enumerate(df.iterrows(), start + 1):
        dep = depreciation_for_row(row, metal_rate)
        taxable = float(row.get("Taxable Amount (₹)", 0) or 0)
        dep_amt = taxable * dep / 100 if row.get("Select", False) else 0
        net = taxable - dep_amt if row.get("Select", False) else 0
        
        vals = [
            bool(row.get("Select", False)), row.get("Item Type", ""), row.get("Description", ""),
            row.get("Part No. / Labour Code", ""), row.get("HSN / SAC", ""), row.get("PMG", ""),
            row.get("Qty", 0), row.get("Unit", ""), row.get("Rate (₹)", 0),
            taxable, row.get("GST %", 0), row.get("GST Amount (₹)", 0), dep, dep_amt, net
        ]
        
        for c, v in enumerate(vals, 1): 
            ws.cell(i, c, v)
            
    for col in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
        
    ws.freeze_panes = f"A{start+1}"
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return out.getvalue()